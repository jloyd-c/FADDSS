"""
Profiling App — Service Layer (Phase 2)
══════════════════════════════════════════════════════════════════════════════

RULE: Views should NEVER write ORM code directly.
      All database mutations flow through one of these service classes.

SERVICE CLASSES
───────────────
  HouseholdService    — CRUD + status transitions + year-over-year comparison
  NormalizationService — NormalizedData population and batch rebuild
  QueryService         — Cross-year concept queries and demographic summaries
  ReportService        — CSV and Excel export (PDF stubbed)

CHANGE LOGGING
──────────────
Every mutation calls HouseholdChangeLog.log_change() with:
  - The top-level Household (for quick household history)
  - target_type/target_id  (which record changed)
  - action                 (CREATED / UPDATED / DELETED / RESTORED / ...)
  - changed_fields         ({"field": {"old": ..., "new": ...}})
  - changed_by             (request.user)
  - ip_address             (request.META.get('REMOTE_ADDR'))

Do NOT put log_change calls in model.save() or signals — only here, because
signals don't have access to request.user.
"""

import csv
import io
import json
import logging
from collections import Counter
from dataclasses import dataclass
from datetime import date
from typing import Iterator

from django.db import transaction
from django.db.models import Count, Max, Prefetch, Q
from django.db.models.functions import ExtractYear
from django.utils import timezone

from .models import (
    Family, FieldMapping, FormSchema, Household,
    HouseholdChangeLog, HouseholdSurvey, NormalizedData,
    Person, ProgramAvailed,
)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Custom Exceptions
# ─────────────────────────────────────────────────────────────────────────────

class ProfilingError(Exception):
    """Base class for all profiling service errors."""


class SurveyAlreadyExistsError(ProfilingError):
    """Raised when a survey for this household+year already exists."""


class InvalidStatusTransitionError(ProfilingError):
    """Raised when a status transition is not allowed."""


class SurveyImmutableError(ProfilingError):
    """Raised when attempting to modify a VERIFIED survey."""


# ─────────────────────────────────────────────────────────────────────────────
# Private Utilities
# ─────────────────────────────────────────────────────────────────────────────

# Status transitions: action → {allowed_from_statuses, to_status}
_TRANSITIONS = {
    'submit':           {'from': {'DRAFT', 'REVISION'}, 'to': 'SUBMITTED'},
    'verify':           {'from': {'SUBMITTED'},          'to': 'VERIFIED'},
    'request_revision': {'from': {'SUBMITTED'},          'to': 'REVISION'},
}

# Fixed Person fields that can be updated via update_person()
_PERSON_FIXED_FIELDS = (
    'role', 'first_name', 'middle_name', 'last_name', 'suffix',
    'date_of_birth', 'age_at_survey', 'gender', 'civil_status',
    'educational_attainment', 'is_registered_voter', 'sectors',
)

# Fixed Family fields that can be updated via update_family()
_FAMILY_FIXED_FIELDS = ('monthly_income_bracket',)


def _diff_dicts(old: dict, new: dict) -> dict:
    """
    Return {key: {"old": ..., "new": ...}} for every key that changed.

    Handles keys present in only one dict (treated as None in the other).

    Example:
        _diff_dicts({'a': 1, 'b': 2}, {'a': 1, 'b': 3, 'c': 4})
        → {'b': {'old': 2, 'new': 3}, 'c': {'old': None, 'new': 4}}
    """
    changes = {}
    for key in set(old) | set(new):
        old_val = old.get(key)
        new_val = new.get(key)
        if old_val != new_val:
            changes[key] = {'old': old_val, 'new': new_val}
    return changes


def _person_match_key(person: Person) -> tuple:
    """
    Stable identity key for matching the same person across two survey years.
    We match on normalised name + date_of_birth because Person has no
    persistent cross-survey ID — each survey creates fresh Person rows.
    """
    return (
        person.last_name.strip().upper(),
        person.first_name.strip().upper(),
        str(person.date_of_birth) if person.date_of_birth else '',
    )


def _apply_transition(survey: HouseholdSurvey, action: str) -> str:
    """
    Validate and return the new status for a survey status transition.

    Raises InvalidStatusTransitionError if the current status is not
    in the allowed set for this action.
    """
    rule = _TRANSITIONS[action]
    if survey.status not in rule['from']:
        allowed = ', '.join(sorted(rule['from']))
        raise InvalidStatusTransitionError(
            f"Cannot '{action}' a survey with status '{survey.status}'. "
            f"Allowed statuses: {allowed}."
        )
    return rule['to']


# ─────────────────────────────────────────────────────────────────────────────
# HouseholdService
# ─────────────────────────────────────────────────────────────────────────────

class HouseholdService:
    """
    All mutations to Household, HouseholdSurvey, Family, Person, and
    ProgramAvailed go through this class.

    IMPORTANT: Every write method logs to HouseholdChangeLog. Do not skip
    the log_change calls — they are required for government compliance.
    """

    # ── Household ────────────────────────────────────────────────────────────

    @staticmethod
    def create_household(
        purok,
        address: str,
        created_by,
        household_number: str | None = None,
        latitude=None,
        longitude=None,
        notes: str = '',
    ) -> Household:
        """
        Create a new Household (physical dwelling) record.

        If household_number is not provided, auto-generates one using
        apps.common.utils.generate_household_number().

        Args:
            purok:            Purok instance
            address:          Full street address
            created_by:       User performing the action
            household_number: Optional explicit number (e.g. "PRK3-2024-001")
            latitude/longitude: Optional GPS coordinates
            notes:            Internal notes

        Returns:
            New Household instance
        """
        if not household_number:
            from apps.common.utils import generate_household_number
            household_number = generate_household_number()

        return Household.objects.create(
            household_number=household_number,
            purok=purok,
            address=address,
            latitude=latitude,
            longitude=longitude,
            notes=notes,
            created_by=created_by,
            updated_by=created_by,
        )

    # ── Survey: full nested create ────────────────────────────────────────────

    @staticmethod
    @transaction.atomic
    def create_survey(
        household: Household,
        form_schema: FormSchema,
        survey_year: int,
        survey_data: dict,
        families_data: list,
        surveyed_by,
        created_by,
        surveyed_at=None,
        ip_address: str | None = None,
    ) -> HouseholdSurvey:
        """
        Create a complete survey (Survey + Families + Persons + Programs)
        in a single atomic transaction.

        families_data format:
        [
          {
            "monthly_income_bracket": "5K_10K",   # optional, defaults to UNSPECIFIED
            "data": {"housing_tenure": "owned"},   # family-level JSON answers
            "persons": [
              {
                "role":                    "HEAD",
                "first_name":              "Juan",
                "middle_name":             "Santos",
                "last_name":               "Dela Cruz",
                "suffix":                  "",
                "date_of_birth":           "1980-03-15",   # optional
                "age_at_survey":           44,             # optional
                "gender":                  "MALE",
                "civil_status":            "MARRIED",
                "educational_attainment":  "COLLEGE_GRAD",
                "is_registered_voter":     True,
                "sectors":                 ["4PS"],
                "data":                    {"occupation": "Farmer"},
              }
            ],
            "programs": [
              {
                "program_type":      "4PS",
                "program_name":      "Pantawid Pamilya",
                "date_availed":      "2024-01-15",
                "amount":            null,
                "reference_no":      "4PS-2024-001",
                "description":       "",
                "data":              {},
                "beneficiary_index": null   # 0-based index into persons list;
                                            # null = whole-family benefit
              }
            ]
          }
        ]

        The transaction guarantees:
        - All families + persons + programs are created or NONE are (atomicity)
        - NormalizedData is populated AFTER the transaction via on_commit()
          (signals.py wires this up automatically)
        - One change log entry is created for the survey creation

        Raises:
            SurveyAlreadyExistsError: if a survey for household+year exists
        """
        # Guard: unique_together constraint
        if HouseholdSurvey.objects.filter(
            household=household,
            survey_year=survey_year,
        ).exists():
            raise SurveyAlreadyExistsError(
                f"A survey for household '{household.household_number}' "
                f"in year {survey_year} already exists."
            )

        # Create the survey
        survey = HouseholdSurvey.objects.create(
            household=household,
            form_schema=form_schema,
            survey_year=survey_year,
            surveyed_by=surveyed_by,
            surveyed_at=surveyed_at,
            status=HouseholdSurvey.SurveyStatus.DRAFT,
            data=survey_data or {},
            created_by=created_by,
            updated_by=created_by,
        )

        # Create families, persons, programs
        for family_dict in families_data:
            HouseholdService._create_family_with_members(
                survey=survey,
                family_dict=family_dict,
                created_by=created_by,
            )

        # Log the creation
        HouseholdChangeLog.log_change(
            household=household,
            target_type=HouseholdChangeLog.TargetType.SURVEY,
            target_id=survey.id,
            action=HouseholdChangeLog.Action.CREATED,
            changed_fields={'survey_year': {'old': None, 'new': survey_year}},
            changed_by=created_by,
            ip_address=ip_address,
            survey_year=survey_year,
        )

        return survey

    @staticmethod
    def _create_family_with_members(survey: HouseholdSurvey, family_dict: dict, created_by):
        """
        Internal helper: create one Family with its Persons and Programs.
        Must be called inside an active transaction.atomic() block.
        """
        # Auto-assign sequential family_number within this survey.
        # MAX + 1 inside the transaction prevents race conditions.
        last_num = Family.objects.filter(
            household_survey=survey
        ).aggregate(m=Max('family_number'))['m'] or 0
        family_number = last_num + 1

        family = Family.objects.create(
            household_survey=survey,
            family_number=family_dict.get('family_number', family_number),
            monthly_income_bracket=family_dict.get(
                'monthly_income_bracket',
                Family.IncomeBracket.UNSPECIFIED,
            ),
            data=family_dict.get('data', {}),
            created_by=created_by,
            updated_by=created_by,
        )

        # Create persons — collect in list to resolve beneficiary_index later
        created_persons: list[Person] = []
        for person_dict in family_dict.get('persons', []):
            person = Person.objects.create(
                family=family,
                role=person_dict.get('role', Person.Role.NON_RELATIVE),
                first_name=person_dict.get('first_name', ''),
                middle_name=person_dict.get('middle_name', ''),
                last_name=person_dict.get('last_name', ''),
                suffix=person_dict.get('suffix', ''),
                date_of_birth=person_dict.get('date_of_birth'),
                age_at_survey=person_dict.get('age_at_survey'),
                gender=person_dict.get('gender', ''),
                civil_status=person_dict.get('civil_status', ''),
                educational_attainment=person_dict.get('educational_attainment', ''),
                is_registered_voter=person_dict.get('is_registered_voter'),
                sectors=person_dict.get('sectors', []),
                data=person_dict.get('data', {}),
                created_by=created_by,
                updated_by=created_by,
            )
            created_persons.append(person)

        # Create programs, resolving beneficiary_index → Person FK
        for prog_dict in family_dict.get('programs', []):
            beneficiary = None
            bindex = prog_dict.get('beneficiary_index')
            if bindex is not None and 0 <= bindex < len(created_persons):
                beneficiary = created_persons[bindex]

            ProgramAvailed.objects.create(
                family=family,
                beneficiary=beneficiary,
                program_type=prog_dict['program_type'],
                program_name=prog_dict.get('program_name', ''),
                date_availed=prog_dict.get('date_availed'),
                amount=prog_dict.get('amount'),
                reference_no=prog_dict.get('reference_no', ''),
                description=prog_dict.get('description', ''),
                data=prog_dict.get('data', {}),
                created_by=created_by,
                updated_by=created_by,
            )

    # ── Survey: data mutations ────────────────────────────────────────────────

    @staticmethod
    @transaction.atomic
    def update_survey_data(
        survey: HouseholdSurvey,
        new_data: dict,
        updated_by,
        ip_address: str | None = None,
    ) -> HouseholdSurvey:
        """
        Update household-level survey data (the JSON `data` field).

        Guards:
          - VERIFIED surveys are immutable — raises SurveyImmutableError
          - Only the `data` JSON is updated here; use update_family() / update_person()
            for family and person data.

        Logs the field-level diff to HouseholdChangeLog.
        """
        if survey.status == HouseholdSurvey.SurveyStatus.VERIFIED:
            raise SurveyImmutableError(
                "Cannot edit a VERIFIED survey. "
                "Use request_revision() to send it back for corrections."
            )

        old_data = dict(survey.data)
        changed_fields = _diff_dicts(old_data, new_data)

        if not changed_fields:
            return survey  # Nothing changed

        survey.data = new_data
        survey.updated_by = updated_by
        survey.save(update_fields=['data', 'updated_by', 'updated_at'])

        HouseholdChangeLog.log_change(
            household=survey.household,
            target_type=HouseholdChangeLog.TargetType.SURVEY,
            target_id=survey.id,
            action=HouseholdChangeLog.Action.UPDATED,
            changed_fields=changed_fields,
            changed_by=updated_by,
            ip_address=ip_address,
            survey_year=survey.survey_year,
        )
        return survey

    @staticmethod
    @transaction.atomic
    def update_family(
        family: Family,
        updates: dict,
        updated_by,
        ip_address: str | None = None,
    ) -> Family:
        """
        Update family fixed fields and/or JSON data.

        `updates` may contain:
          - 'monthly_income_bracket': new bracket value
          - 'data': new family-level JSON answers

        The method diffs both and logs a single change entry.
        """
        survey = family.household_survey
        if survey.status == HouseholdSurvey.SurveyStatus.VERIFIED:
            raise SurveyImmutableError("Cannot edit a family in a VERIFIED survey.")

        changed_fields = {}

        # Fixed field changes
        for field in _FAMILY_FIXED_FIELDS:
            if field in updates:
                old_val = getattr(family, field)
                new_val = updates[field]
                if old_val != new_val:
                    changed_fields[field] = {'old': old_val, 'new': new_val}
                    setattr(family, field, new_val)

        # JSON data changes
        if 'data' in updates:
            old_data = dict(family.data)
            new_data = updates['data']
            data_diff = _diff_dicts(old_data, new_data)
            if data_diff:
                changed_fields['data'] = {'old': old_data, 'new': new_data}
            family.data = new_data

        if not changed_fields:
            return family

        family.updated_by = updated_by
        family.save()

        HouseholdChangeLog.log_change(
            household=survey.household,
            target_type=HouseholdChangeLog.TargetType.FAMILY,
            target_id=family.id,
            action=HouseholdChangeLog.Action.UPDATED,
            changed_fields=changed_fields,
            changed_by=updated_by,
            ip_address=ip_address,
            survey_year=survey.survey_year,
        )
        return family

    @staticmethod
    @transaction.atomic
    def update_person(
        person: Person,
        updates: dict,
        updated_by,
        ip_address: str | None = None,
    ) -> Person:
        """
        Update person fixed fields and/or JSON data.

        `updates` may contain any of the _PERSON_FIXED_FIELDS keys,
        plus 'data' for the JSON field.

        NOTE: Pass person with select_related('family__household_survey__household')
              to avoid N+1 queries in the log_change call.
        """
        survey = person.family.household_survey
        if survey.status == HouseholdSurvey.SurveyStatus.VERIFIED:
            raise SurveyImmutableError("Cannot edit a person in a VERIFIED survey.")

        changed_fields = {}

        for field in _PERSON_FIXED_FIELDS:
            if field in updates:
                old_val = getattr(person, field)
                new_val = updates[field]
                if old_val != new_val:
                    changed_fields[field] = {'old': old_val, 'new': new_val}
                    setattr(person, field, new_val)

        if 'data' in updates:
            old_data = dict(person.data)
            new_data = updates['data']
            data_diff = _diff_dicts(old_data, new_data)
            if data_diff:
                changed_fields['data'] = {'old': old_data, 'new': new_data}
            person.data = new_data

        if not changed_fields:
            return person

        person.updated_by = updated_by
        person.save()

        HouseholdChangeLog.log_change(
            household=survey.household,
            target_type=HouseholdChangeLog.TargetType.PERSON,
            target_id=person.id,
            action=HouseholdChangeLog.Action.UPDATED,
            changed_fields=changed_fields,
            changed_by=updated_by,
            ip_address=ip_address,
            survey_year=survey.survey_year,
        )
        return person

    # ── Survey: status transitions ────────────────────────────────────────────

    @staticmethod
    @transaction.atomic
    def submit_survey(
        survey: HouseholdSurvey,
        submitted_by,
        ip_address: str | None = None,
    ) -> HouseholdSurvey:
        """
        DRAFT|REVISION → SUBMITTED.
        Called by the data-entry staff when they believe the survey is complete.
        """
        new_status = _apply_transition(survey, 'submit')

        survey.status = new_status
        survey.updated_by = submitted_by
        survey.save(update_fields=['status', 'updated_by', 'updated_at'])

        HouseholdChangeLog.log_change(
            household=survey.household,
            target_type=HouseholdChangeLog.TargetType.SURVEY,
            target_id=survey.id,
            action=HouseholdChangeLog.Action.UPDATED,
            changed_fields={'status': {'old': survey.status, 'new': new_status}},
            changed_by=submitted_by,
            ip_address=ip_address,
            survey_year=survey.survey_year,
        )
        return survey

    @staticmethod
    @transaction.atomic
    def verify_survey(
        survey: HouseholdSurvey,
        verified_by,
        notes: str = '',
        ip_address: str | None = None,
    ) -> HouseholdSurvey:
        """
        SUBMITTED → VERIFIED.
        Called by a supervisor/admin after reviewing the survey.
        Sets verified_by and verified_at.
        """
        new_status = _apply_transition(survey, 'verify')
        old_status = survey.status

        survey.status = new_status
        survey.verified_by = verified_by
        survey.verified_at = timezone.now()
        if notes:
            survey.notes = notes
        survey.updated_by = verified_by
        survey.save(update_fields=[
            'status', 'verified_by', 'verified_at', 'notes', 'updated_by', 'updated_at'
        ])

        HouseholdChangeLog.log_change(
            household=survey.household,
            target_type=HouseholdChangeLog.TargetType.SURVEY,
            target_id=survey.id,
            action=HouseholdChangeLog.Action.VERIFIED,
            changed_fields={'status': {'old': old_status, 'new': new_status}},
            changed_by=verified_by,
            ip_address=ip_address,
            survey_year=survey.survey_year,
            notes=notes,
        )
        return survey

    @staticmethod
    @transaction.atomic
    def request_revision(
        survey: HouseholdSurvey,
        requested_by,
        notes: str = '',
        ip_address: str | None = None,
    ) -> HouseholdSurvey:
        """
        SUBMITTED → REVISION.
        Called by a supervisor/admin when the survey needs corrections.
        """
        new_status = _apply_transition(survey, 'request_revision')
        old_status = survey.status

        survey.status = new_status
        survey.notes = notes
        survey.updated_by = requested_by
        survey.save(update_fields=['status', 'notes', 'updated_by', 'updated_at'])

        HouseholdChangeLog.log_change(
            household=survey.household,
            target_type=HouseholdChangeLog.TargetType.SURVEY,
            target_id=survey.id,
            action=HouseholdChangeLog.Action.REVISION,
            changed_fields={'status': {'old': old_status, 'new': new_status}},
            changed_by=requested_by,
            ip_address=ip_address,
            survey_year=survey.survey_year,
            notes=notes,
        )
        return survey

    @staticmethod
    @transaction.atomic
    def soft_delete_survey(
        survey: HouseholdSurvey,
        deleted_by,
        ip_address: str | None = None,
    ) -> None:
        """
        Soft-delete a survey and all its families, persons, and programs.
        NEVER calls .delete() — only sets is_deleted=True on all records.

        The cascade is manual (not DB-level) so each deletion is logged.
        """
        # Cascade soft delete to children
        for family in Family.all_objects.filter(household_survey=survey):
            Person.all_objects.filter(family=family).update(
                is_deleted=True, deleted_at=timezone.now(), deleted_by=deleted_by
            )
            ProgramAvailed.all_objects.filter(family=family).update(
                is_deleted=True, deleted_at=timezone.now(), deleted_by=deleted_by
            )
            family.soft_delete(deleted_by_user=deleted_by)

        survey.soft_delete(deleted_by_user=deleted_by)

        HouseholdChangeLog.log_change(
            household=survey.household,
            target_type=HouseholdChangeLog.TargetType.SURVEY,
            target_id=survey.id,
            action=HouseholdChangeLog.Action.DELETED,
            changed_fields={'is_deleted': {'old': False, 'new': True}},
            changed_by=deleted_by,
            ip_address=ip_address,
            survey_year=survey.survey_year,
        )

    # ── Cross-year comparison ─────────────────────────────────────────────────

    @staticmethod
    def compare_surveys(
        survey_a: HouseholdSurvey,
        survey_b: HouseholdSurvey,
    ) -> dict:
        """
        Produce a structured diff between two surveys (any two years) for the
        same or different households.

        Household-level diff uses NormalizedData (canonical names) so that
        2024 "water_source"="metered" and 2026 "water_access_level"="level_3"
        are correctly identified as the same concept.

        Family and person level diffs use raw field values directly, with
        family matching by family_number and person matching by name+DOB.

        Returns:
        {
          "survey_a": {id, year, status},
          "survey_b": {id, year, status},
          "normalized_data_warning": bool,   # True if NormalizedData is missing
          "household_diff": {
            "canonical_name": {
              "a": {"raw": "...", "canonical": "..."},
              "b": {"raw": "...", "canonical": "..."}
            }, ...
          },
          "families": [
            {
              "family_number": 1,
              "match": "both" | "only_in_a" | "only_in_b",
              "family_a_id": "...",
              "family_b_id": "...",
              "family_diff": {"monthly_income_bracket": {"a": ..., "b": ...}},
              "persons": [
                {
                  "match": "both" | "only_in_a" | "only_in_b",
                  "name": "...",
                  "person_a_id": "...",
                  "person_b_id": "...",
                  "person_diff": {field: {"a": ..., "b": ...}}
                }
              ]
            }
          ],
          "summary": {
            "household_fields_changed": int,
            "families_added": int,
            "families_removed": int,
            "families_changed": int,
            "persons_added": int,
            "persons_removed": int,
          }
        }

        Example:
            survey_2024 = HouseholdSurvey.objects.get(household=hh, survey_year=2024)
            survey_2026 = HouseholdSurvey.objects.get(household=hh, survey_year=2026)
            diff = HouseholdService.compare_surveys(survey_2024, survey_2026)
        """
        # ── Household-level diff via NormalizedData ──────────────────────────
        nd_a = {nd.canonical_name: nd for nd in
                NormalizedData.objects.filter(household_survey=survey_a, level='household')}
        nd_b = {nd.canonical_name: nd for nd in
                NormalizedData.objects.filter(household_survey=survey_b, level='household')}

        # Warn if NormalizedData hasn't been populated yet
        missing_nd = not nd_a and bool(survey_a.data)

        household_diff = {}
        for concept in sorted(set(nd_a) | set(nd_b)):
            val_a = nd_a[concept].canonical_value if concept in nd_a else None
            val_b = nd_b[concept].canonical_value if concept in nd_b else None
            if val_a != val_b:
                household_diff[concept] = {
                    'a': {
                        'raw':       nd_a[concept].raw_value if concept in nd_a else None,
                        'canonical': val_a,
                    },
                    'b': {
                        'raw':       nd_b[concept].raw_value if concept in nd_b else None,
                        'canonical': val_b,
                    },
                }

        # ── Family comparison ─────────────────────────────────────────────────
        fams_a = {
            f.family_number: f
            for f in Family.objects.prefetch_related(
                Prefetch('persons', queryset=Person.objects.all())
            ).filter(household_survey=survey_a)
        }
        fams_b = {
            f.family_number: f
            for f in Family.objects.prefetch_related(
                Prefetch('persons', queryset=Person.objects.all())
            ).filter(household_survey=survey_b)
        }

        families_result = []
        summary = {
            'household_fields_changed': len(household_diff),
            'families_added': 0,
            'families_removed': 0,
            'families_changed': 0,
            'persons_added': 0,
            'persons_removed': 0,
        }

        for fnum in sorted(set(fams_a) | set(fams_b)):
            fa = fams_a.get(fnum)
            fb = fams_b.get(fnum)

            if fa and not fb:
                summary['families_removed'] += 1
                families_result.append({
                    'family_number': fnum,
                    'match': 'only_in_a',
                    'family_a_id': str(fa.id),
                    'family_b_id': None,
                    'family_diff': {},
                    'persons': [],
                })
                continue

            if fb and not fa:
                summary['families_added'] += 1
                families_result.append({
                    'family_number': fnum,
                    'match': 'only_in_b',
                    'family_a_id': None,
                    'family_b_id': str(fb.id),
                    'family_diff': {},
                    'persons': [],
                })
                continue

            # Both families exist — compare fixed fields and persons
            family_diff = {}
            if fa.monthly_income_bracket != fb.monthly_income_bracket:
                family_diff['monthly_income_bracket'] = {
                    'a': fa.monthly_income_bracket,
                    'b': fb.monthly_income_bracket,
                }

            persons_result, p_added, p_removed = _compare_persons(
                list(fa.persons.all()), list(fb.persons.all())
            )
            summary['persons_added']   += p_added
            summary['persons_removed'] += p_removed

            has_changes = bool(family_diff or p_added or p_removed or
                               any(pr['person_diff'] for pr in persons_result))
            if has_changes:
                summary['families_changed'] += 1

            families_result.append({
                'family_number': fnum,
                'match': 'both',
                'family_a_id': str(fa.id),
                'family_b_id': str(fb.id),
                'family_diff': family_diff,
                'persons': persons_result,
            })

        return {
            'survey_a': {
                'id': str(survey_a.id),
                'year': survey_a.survey_year,
                'status': survey_a.status,
            },
            'survey_b': {
                'id': str(survey_b.id),
                'year': survey_b.survey_year,
                'status': survey_b.status,
            },
            'normalized_data_warning': missing_nd,
            'household_diff': household_diff,
            'families': families_result,
            'summary': summary,
        }

    @staticmethod
    def get_latest_survey(household: Household) -> HouseholdSurvey | None:
        """
        Return the most recent non-deleted HouseholdSurvey for this household,
        or None if no surveys have been taken.
        """
        return (
            HouseholdSurvey.objects
            .filter(household=household)
            .select_related('form_schema', 'surveyed_by')
            .order_by('-survey_year')
            .first()
        )


def _compare_persons(persons_a: list, persons_b: list) -> tuple[list, int, int]:
    """
    Match persons across two family snapshots and produce per-person diffs.

    Returns:
        (persons_result, added_count, removed_count)
    """
    COMPARABLE_FIELDS = (
        'role', 'gender', 'civil_status', 'educational_attainment',
        'is_registered_voter', 'sectors',
    )
    map_a = {_person_match_key(p): p for p in persons_a}
    map_b = {_person_match_key(p): p for p in persons_b}

    results = []
    added = removed = 0

    for key in set(map_a) | set(map_b):
        pa = map_a.get(key)
        pb = map_b.get(key)
        display_name = (pa or pb).full_name

        if pa and not pb:
            removed += 1
            results.append({
                'match': 'only_in_a',
                'name': display_name,
                'person_a_id': str(pa.id),
                'person_b_id': None,
                'person_diff': {},
            })
        elif pb and not pa:
            added += 1
            results.append({
                'match': 'only_in_b',
                'name': display_name,
                'person_a_id': None,
                'person_b_id': str(pb.id),
                'person_diff': {},
            })
        else:
            diff = {}
            for field in COMPARABLE_FIELDS:
                va, vb = getattr(pa, field), getattr(pb, field)
                if va != vb:
                    diff[field] = {'a': va, 'b': vb}
            results.append({
                'match': 'both',
                'name': display_name,
                'person_a_id': str(pa.id),
                'person_b_id': str(pb.id),
                'person_diff': diff,
            })

    return results, added, removed


# ─────────────────────────────────────────────────────────────────────────────
# NormalizationService
# ─────────────────────────────────────────────────────────────────────────────

class NormalizationService:
    """
    Manages the NormalizedData table.

    NOTE: NormalizedData is auto-populated by signal handlers in signals.py
    whenever a HouseholdSurvey, Family, or Person is saved. Use these methods
    for:
      - Manual refresh after correcting a FieldMapping
      - Backfill after adding a new FieldMapping
      - Admin/management command batch rebuilds
    """

    @staticmethod
    def populate_normalized_data(survey: HouseholdSurvey) -> dict:
        """
        Fully regenerate NormalizedData for one survey and all its children.

        This is a synchronous, direct call — no transaction.on_commit deferral.
        Safe to call from management commands and admin actions.

        Returns:
            {'household': N, 'families': N, 'persons': N}
        """
        from .signals import rebuild_normalized_data_for_survey
        return rebuild_normalized_data_for_survey(survey)

    @staticmethod
    def rebuild_all_normalized_data(
        year: int | None = None,
        batch_size: int = 50,
    ) -> dict:
        """
        Batch rebuild NormalizedData for all surveys (or all surveys in one year).

        Processes surveys in batches to avoid loading the entire table into memory.
        Each survey is rebuilt independently — one failure doesn't stop the rest.

        Args:
            year:       Rebuild only this survey year. None = rebuild everything.
            batch_size: Number of surveys to fetch per DB round-trip.

        Returns:
            {
                'surveys_processed': int,
                'total_rows_inserted': int,
                'errors': [{'survey_id': '...', 'error': '...'}]
            }
        """
        from .signals import rebuild_normalized_data_for_survey

        qs = (
            HouseholdSurvey.all_objects
            .select_related('form_schema')
            .prefetch_related(
                Prefetch('families', queryset=Family.all_objects.all()),
                Prefetch(
                    'families__persons',
                    queryset=Person.all_objects.all(),
                ),
            )
        )
        if year is not None:
            qs = qs.filter(survey_year=year)

        processed = 0
        total_rows = 0
        errors = []

        for survey in qs.iterator(chunk_size=batch_size):
            try:
                counts = rebuild_normalized_data_for_survey(survey)
                total_rows += sum(counts.values())
                processed += 1
            except Exception as exc:
                errors.append({'survey_id': str(survey.pk), 'error': str(exc)})
                logger.exception(
                    '[NormalizationService] Failed to rebuild survey %s', survey.pk
                )

        logger.info(
            '[NormalizationService] Rebuild complete: %d surveys, %d rows, %d errors',
            processed, total_rows, len(errors),
        )
        return {
            'surveys_processed': processed,
            'total_rows_inserted': total_rows,
            'errors': errors,
        }


# ─────────────────────────────────────────────────────────────────────────────
# QueryService
# ─────────────────────────────────────────────────────────────────────────────

class QueryService:
    """
    Cross-year queries using the NormalizedData table.

    All query methods return querysets or plain dicts — never raw SQL strings.
    """

    @staticmethod
    def filter_by_concept(
        canonical_name: str,
        canonical_value: str,
        year_start: int,
        year_end: int,
        level: str = 'household',
        purok_ids: list | None = None,
    ):
        """
        Find all HouseholdSurveys where a canonical concept has a given value
        within a range of survey years.

        Uses the NormalizedData table, so field names that changed across years
        are automatically handled — one query covers all years.

        Args:
            canonical_name:  e.g. "water_source"
            canonical_value: e.g. "level_3"   (canonical, not raw)
            year_start:      inclusive start year
            year_end:        inclusive end year
            level:           'household'|'family'|'person'
            purok_ids:       Optional list of Purok PKs to restrict results

        Returns:
            HouseholdSurvey queryset

        Example:
            # Find all households with metered water (2024-2026)
            QueryService.filter_by_concept(
                canonical_name='water_source',
                canonical_value='level_3',
                year_start=2024,
                year_end=2026,
            )
        """
        matching_survey_ids = (
            NormalizedData.objects
            .filter(
                canonical_name=canonical_name,
                canonical_value=canonical_value,
                level=level,
                survey_year__gte=year_start,
                survey_year__lte=year_end,
            )
            .values_list('household_survey_id', flat=True)
            .distinct()
        )

        qs = (
            HouseholdSurvey.objects
            .filter(pk__in=matching_survey_ids)
            .select_related('household__purok', 'form_schema')
            .order_by('household__household_number', 'survey_year')
        )

        if purok_ids:
            qs = qs.filter(household__purok_id__in=purok_ids)

        return qs

    @staticmethod
    def get_trend(
        canonical_name: str,
        canonical_value: str,
        years: list[int],
        purok_ids: list | None = None,
    ) -> list[dict]:
        """
        Count how many unique households match a canonical concept+value
        for each year in the given list.

        Fills in zero for years with no matches so the caller always gets
        a complete year series.

        Returns:
            [{'year': 2024, 'count': 45}, {'year': 2025, 'count': 52}, ...]

        Example:
            # How many households upgraded to metered water each year?
            QueryService.get_trend(
                canonical_name='water_source',
                canonical_value='level_3',
                years=[2024, 2025, 2026],
            )
        """
        qs = NormalizedData.objects.filter(
            canonical_name=canonical_name,
            canonical_value=canonical_value,
            survey_year__in=years,
        )
        if purok_ids:
            qs = qs.filter(household_survey__household__purok_id__in=purok_ids)

        rows = (
            qs
            .values('survey_year')
            # distinct=True prevents double-counting when a survey has
            # multiple NormalizedData rows for the same canonical concept
            .annotate(count=Count('household_survey', distinct=True))
            .order_by('survey_year')
        )

        result_map = {row['survey_year']: row['count'] for row in rows}
        return [{'year': y, 'count': result_map.get(y, 0)} for y in sorted(years)]

    @staticmethod
    def search_persons(
        query: str,
        purok_ids: list | None = None,
        survey_year: int | None = None,
        sectors: list | None = None,
        gender: str | None = None,
        limit: int = 50,
    ):
        """
        Full-name search for persons, with optional demographic filters.

        Args:
            query:       Partial name (searches first_name, last_name, middle_name)
            purok_ids:   Restrict to persons in these puroks
            survey_year: Restrict to a specific survey year
            sectors:     List of sector codes — returns persons with ALL listed sectors
            gender:      'MALE' | 'FEMALE' | 'OTHER'
            limit:       Maximum results to return

        Returns:
            Person queryset with select_related to household and survey

        Example:
            # Find all PWD persons named "Juan" surveyed in 2024
            QueryService.search_persons(query='Juan', survey_year=2024, sectors=['PWD'])
        """
        qs = (
            Person.objects
            .select_related(
                'family__household_survey__household__purok',
                'family__household_survey',
            )
        )

        if query:
            qs = qs.filter(
                Q(first_name__icontains=query)
                | Q(last_name__icontains=query)
                | Q(middle_name__icontains=query)
            )

        if survey_year is not None:
            qs = qs.filter(family__household_survey__survey_year=survey_year)

        if purok_ids:
            qs = qs.filter(
                family__household_survey__household__purok_id__in=purok_ids
            )

        if sectors:
            # JSONField containment: person must have ALL specified sectors
            for sector in sectors:
                qs = qs.filter(sectors__contains=[sector])

        if gender:
            qs = qs.filter(gender=gender)

        return qs.order_by('last_name', 'first_name')[:limit]

    @staticmethod
    def get_demographics_summary(
        survey_year: int,
        purok_ids: list | None = None,
    ) -> dict:
        """
        Aggregate demographic statistics for a given survey year.

        Used for dashboard widgets and annual reports.

        Returns a dict with counts broken down by gender, age group,
        civil status, educational attainment, sector, and income bracket.

        NOTE: Sector counts use Python-side aggregation (Counter) because
        PostgreSQL cannot GROUP BY elements of a JSONB array without a raw
        SQL unnest. At typical barangay scale (< 20,000 persons) this is fast.

        Example:
            summary = QueryService.get_demographics_summary(2024, purok_ids=[1,2])
        """
        base_qs = HouseholdSurvey.objects.filter(
            survey_year=survey_year, is_deleted=False
        )
        if purok_ids:
            base_qs = base_qs.filter(household__purok_id__in=purok_ids)

        survey_ids = base_qs.values_list('id', flat=True)

        # Aggregated counts
        household_count = base_qs.count()
        family_count    = Family.objects.filter(
            household_survey_id__in=survey_ids, is_deleted=False
        ).count()

        persons_qs = Person.objects.filter(
            family__household_survey_id__in=survey_ids,
            is_deleted=False,
        )
        person_count = persons_qs.count()

        # Gender breakdown (pure ORM)
        gender_counts = dict(
            persons_qs.values('gender')
            .annotate(count=Count('id'))
            .values_list('gender', 'count')
        )

        # Civil status breakdown
        civil_status_counts = dict(
            persons_qs.values('civil_status')
            .annotate(count=Count('id'))
            .values_list('civil_status', 'count')
        )

        # Educational attainment breakdown
        education_counts = dict(
            persons_qs.values('educational_attainment')
            .annotate(count=Count('id'))
            .values_list('educational_attainment', 'count')
        )

        # Age groups — derive from date_of_birth vs survey_year
        age_qs = persons_qs.exclude(date_of_birth=None).annotate(
            survey_age=survey_year - ExtractYear('date_of_birth')
        )
        age_groups = {
            'child_0_11':  age_qs.filter(survey_age__lte=11).count(),
            'youth_12_17': age_qs.filter(survey_age__range=(12, 17)).count(),
            'adult_18_59': age_qs.filter(survey_age__range=(18, 59)).count(),
            'senior_60_plus': age_qs.filter(survey_age__gte=60).count(),
            'unknown':     persons_qs.filter(date_of_birth=None).count(),
        }

        # Sector counts — Python-side aggregation (see docstring)
        sector_counter: Counter = Counter()
        for sectors_list in persons_qs.values_list('sectors', flat=True):
            for sector in (sectors_list or []):
                sector_counter[sector] += 1

        # Income bracket breakdown (Family level)
        income_counts = dict(
            Family.objects.filter(
                household_survey_id__in=survey_ids, is_deleted=False
            )
            .values('monthly_income_bracket')
            .annotate(count=Count('id'))
            .values_list('monthly_income_bracket', 'count')
        )

        return {
            'survey_year':          survey_year,
            'total_households':     household_count,
            'total_families':       family_count,
            'total_persons':        person_count,
            'gender':               gender_counts,
            'civil_status':         civil_status_counts,
            'educational_attainment': education_counts,
            'age_groups':           age_groups,
            'sectors':              dict(sector_counter),
            'income_brackets':      income_counts,
        }


# ─────────────────────────────────────────────────────────────────────────────
# ReportService
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ExportResult:
    """Returned by ReportService.generate_export(). Pass directly to HttpResponse."""
    content:      bytes
    content_type: str
    filename:     str


class ReportService:
    """
    Generates downloadable exports (CSV, Excel) from profiling data.

    Supported entity_types:
        'household' — Household records with latest survey summary
        'survey'    — HouseholdSurvey records
        'family'    — Family records with household+survey context
        'person'    — All Person records (demographic data)
        'program'   — ProgramAvailed records

    Supported formats:
        'csv'   — Single-sheet CSV (utf-8-sig BOM for Excel compatibility)
        'excel' — Single-sheet .xlsx (requires openpyxl; add to requirements)
        'pdf'   — Not yet implemented (requires weasyprint or reportlab)

    Filters dict (all optional):
        'survey_year':  int or [int, int] range
        'purok_ids':    list of Purok PKs
        'status':       HouseholdSurvey status filter
        'include_deleted': bool (default False)
    """

    # Column definitions for each entity type
    ENTITY_COLUMNS = {
        'household': [
            'household_number', 'purok', 'address', 'status',
            'latitude', 'longitude', 'total_surveys', 'latest_survey_year',
            'created_at',
        ],
        'survey': [
            'household_number', 'purok', 'survey_year', 'status',
            'surveyed_by', 'surveyed_at', 'verified_by', 'verified_at',
            'family_count', 'person_count', 'form_schema', 'created_at',
        ],
        'family': [
            'household_number', 'purok', 'survey_year', 'family_number',
            'monthly_income_bracket', 'person_count', 'programs_count',
        ],
        'person': [
            'household_number', 'purok', 'survey_year', 'family_number',
            'last_name', 'first_name', 'middle_name', 'suffix',
            'role', 'gender', 'date_of_birth', 'age_at_survey',
            'civil_status', 'educational_attainment',
            'is_registered_voter', 'sectors',
        ],
        'program': [
            'household_number', 'purok', 'survey_year', 'family_number',
            'beneficiary_name', 'program_type', 'program_name',
            'date_availed', 'amount', 'reference_no', 'description',
        ],
    }

    @classmethod
    def generate_export(
        cls,
        entity_type: str,
        filters: dict,
        fmt: str = 'csv',
    ) -> ExportResult:
        """
        Generate a downloadable export file.

        Args:
            entity_type: One of ENTITY_COLUMNS keys
            filters:     Dict of filter params (see class docstring)
            fmt:         'csv' | 'excel' | 'pdf'

        Returns:
            ExportResult with content bytes, content_type, and filename

        Example:
            result = ReportService.generate_export(
                entity_type='person',
                filters={'survey_year': 2024, 'purok_ids': [1, 2]},
                fmt='excel',
            )
            response = HttpResponse(result.content, content_type=result.content_type)
            response['Content-Disposition'] = f'attachment; filename="{result.filename}"'
        """
        if entity_type not in cls.ENTITY_COLUMNS:
            raise ValueError(
                f"Unknown entity_type '{entity_type}'. "
                f"Choose from: {', '.join(cls.ENTITY_COLUMNS)}"
            )

        columns = cls.ENTITY_COLUMNS[entity_type]
        today = date.today().isoformat()

        if fmt == 'csv':
            return cls._export_csv(entity_type, columns, filters, today)
        elif fmt == 'excel':
            return cls._export_excel(entity_type, columns, filters, today)
        elif fmt == 'pdf':
            raise NotImplementedError(
                "PDF export is not yet implemented. "
                "Install 'weasyprint' or 'reportlab' and implement _export_pdf()."
            )
        else:
            raise ValueError(f"Unknown format '{fmt}'. Choose from: csv, excel, pdf")

    @classmethod
    def _export_csv(cls, entity_type, columns, filters, today) -> ExportResult:
        """
        Stream rows into an in-memory CSV buffer.
        Uses utf-8-sig BOM so Filipino characters display correctly in Excel.
        """
        buffer = io.StringIO()
        writer = csv.DictWriter(
            buffer,
            fieldnames=columns,
            extrasaction='ignore',
            lineterminator='\r\n',
        )
        writer.writeheader()
        for row in cls._get_rows(entity_type, filters):
            writer.writerow(row)

        content = buffer.getvalue().encode('utf-8-sig')  # BOM for Excel
        return ExportResult(
            content=content,
            content_type='text/csv; charset=utf-8-sig',
            filename=f'{entity_type}_export_{today}.csv',
        )

    @classmethod
    def _export_excel(cls, entity_type, columns, filters, today) -> ExportResult:
        """
        Build an .xlsx workbook with one styled sheet.
        Requires openpyxl (pip install openpyxl).
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Run: pip install openpyxl"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = entity_type.title()

        # Header row styling
        header_font   = Font(bold=True, color='FFFFFF')
        header_fill   = PatternFill('solid', fgColor='1F4E79')
        header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 20

        for col_idx, col_name in enumerate(columns, start=1):
            cell             = ws.cell(row=1, column=col_idx, value=col_name.replace('_', ' ').title())
            cell.font        = header_font
            cell.fill        = header_fill
            cell.alignment   = header_align

        # Data rows
        for row_idx, row in enumerate(cls._get_rows(entity_type, filters), start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        # Auto-size columns (capped at 60 chars)
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(
                (len(str(cell.value or '')) for cell in col_cells),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        buffer = io.BytesIO()
        wb.save(buffer)
        return ExportResult(
            content=buffer.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
            filename=f'{entity_type}_export_{today}.xlsx',
        )

    # ── Row builders ──────────────────────────────────────────────────────────

    @classmethod
    def _get_rows(cls, entity_type: str, filters: dict) -> Iterator[dict]:
        """Route to the correct row builder based on entity_type."""
        builders = {
            'household': cls._household_rows,
            'survey':    cls._survey_rows,
            'family':    cls._family_rows,
            'person':    cls._person_rows,
            'program':   cls._program_rows,
        }
        yield from builders[entity_type](filters)

    @classmethod
    def _base_survey_filter(cls, qs, filters: dict):
        """Apply common survey-level filters."""
        year = filters.get('survey_year')
        if isinstance(year, int):
            qs = qs.filter(survey_year=year)
        elif isinstance(year, (list, tuple)) and len(year) == 2:
            qs = qs.filter(survey_year__gte=year[0], survey_year__lte=year[1])

        purok_ids = filters.get('purok_ids')
        if purok_ids:
            qs = qs.filter(household__purok_id__in=purok_ids)

        status = filters.get('status')
        if status:
            qs = qs.filter(status=status)

        return qs

    @classmethod
    def _household_rows(cls, filters: dict) -> Iterator[dict]:
        manager = Household.all_objects if filters.get('include_deleted') else Household.objects
        qs = manager.select_related('purok').annotate(
            total_surveys=Count('surveys', distinct=True),
        ).order_by('purok__number', 'household_number')

        if filters.get('purok_ids'):
            qs = qs.filter(purok_id__in=filters['purok_ids'])

        for h in qs.iterator(chunk_size=200):
            latest = h.surveys.order_by('-survey_year').values('survey_year').first()
            yield {
                'household_number':   h.household_number,
                'purok':              str(h.purok),
                'address':            h.address,
                'status':             h.status,
                'latitude':           h.latitude,
                'longitude':          h.longitude,
                'total_surveys':      h.total_surveys,
                'latest_survey_year': latest['survey_year'] if latest else '',
                'created_at':         h.created_at.date().isoformat() if h.created_at else '',
            }

    @classmethod
    def _survey_rows(cls, filters: dict) -> Iterator[dict]:
        manager = HouseholdSurvey.all_objects if filters.get('include_deleted') else HouseholdSurvey.objects
        qs = (
            manager
            .select_related('household__purok', 'form_schema', 'surveyed_by', 'verified_by')
            .annotate(
                family_count=Count('families', distinct=True),
                person_count=Count('families__persons', distinct=True),
            )
        )
        qs = cls._base_survey_filter(qs, filters)

        for s in qs.iterator(chunk_size=200):
            yield {
                'household_number': s.household.household_number,
                'purok':            str(s.household.purok),
                'survey_year':      s.survey_year,
                'status':           s.status,
                'surveyed_by':      s.surveyed_by.full_name if s.surveyed_by else '',
                'surveyed_at':      str(s.surveyed_at or ''),
                'verified_by':      s.verified_by.full_name if s.verified_by else '',
                'verified_at':      s.verified_at.date().isoformat() if s.verified_at else '',
                'family_count':     s.family_count,
                'person_count':     s.person_count,
                'form_schema':      s.form_schema.name,
                'created_at':       s.created_at.date().isoformat() if s.created_at else '',
            }

    @classmethod
    def _family_rows(cls, filters: dict) -> Iterator[dict]:
        manager = Family.all_objects if filters.get('include_deleted') else Family.objects
        qs = (
            manager
            .select_related('household_survey__household__purok')
            .annotate(
                person_count=Count('persons', distinct=True),
                programs_count=Count('programs_availed', distinct=True),
            )
        )
        survey_qs = cls._base_survey_filter(
            HouseholdSurvey.objects, filters
        )
        qs = qs.filter(household_survey__in=survey_qs)

        for f in qs.iterator(chunk_size=200):
            yield {
                'household_number':      f.household_survey.household.household_number,
                'purok':                 str(f.household_survey.household.purok),
                'survey_year':           f.household_survey.survey_year,
                'family_number':         f.family_number,
                'monthly_income_bracket': f.monthly_income_bracket,
                'person_count':          f.person_count,
                'programs_count':        f.programs_count,
            }

    @classmethod
    def _person_rows(cls, filters: dict) -> Iterator[dict]:
        manager = Person.all_objects if filters.get('include_deleted') else Person.objects
        qs = (
            manager
            .select_related(
                'family__household_survey__household__purok',
            )
        )
        survey_qs = cls._base_survey_filter(HouseholdSurvey.objects, filters)
        qs = qs.filter(family__household_survey__in=survey_qs)

        for p in qs.iterator(chunk_size=200):
            survey = p.family.household_survey
            yield {
                'household_number':      survey.household.household_number,
                'purok':                 str(survey.household.purok),
                'survey_year':           survey.survey_year,
                'family_number':         p.family.family_number,
                'last_name':             p.last_name,
                'first_name':            p.first_name,
                'middle_name':           p.middle_name,
                'suffix':                p.suffix,
                'role':                  p.role,
                'gender':                p.gender,
                'date_of_birth':         str(p.date_of_birth or ''),
                'age_at_survey':         p.age_at_survey,
                'civil_status':          p.civil_status,
                'educational_attainment': p.educational_attainment,
                'is_registered_voter':   p.is_registered_voter,
                # Pipe-delimited for clean CSV — avoids brackets/quotes
                'sectors':               '|'.join(p.sectors or []),
            }

    @classmethod
    def _program_rows(cls, filters: dict) -> Iterator[dict]:
        manager = ProgramAvailed.all_objects if filters.get('include_deleted') else ProgramAvailed.objects
        qs = (
            manager
            .select_related(
                'family__household_survey__household__purok',
                'beneficiary',
            )
        )
        survey_qs = cls._base_survey_filter(HouseholdSurvey.objects, filters)
        qs = qs.filter(family__household_survey__in=survey_qs)

        for prog in qs.iterator(chunk_size=200):
            survey = prog.family.household_survey
            yield {
                'household_number': survey.household.household_number,
                'purok':            str(survey.household.purok),
                'survey_year':      survey.survey_year,
                'family_number':    prog.family.family_number,
                'beneficiary_name': prog.beneficiary.full_name if prog.beneficiary else '',
                'program_type':     prog.program_type,
                'program_name':     prog.program_name,
                'date_availed':     str(prog.date_availed or ''),
                'amount':           str(prog.amount or ''),
                'reference_no':     prog.reference_no,
                'description':      prog.description,
            }
